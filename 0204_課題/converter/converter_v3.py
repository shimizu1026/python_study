"""
Excel変換ロジック
入力ファイルを読み込み、指定のフォーマットに変換する
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelConverter:
    """Excelデータ変換クラス"""
    
    def __init__(self):
        """初期化"""
        self.input_data = None
        self.output_wb = None
        
    def convert(self, input_path, output_path):
        """
        メイン変換処理
        
        Parameters:
        -----------
        input_path : str
            入力ファイルのパス
        output_path : str
            出力ファイルのパス
        """
        # 1. 入力ファイルを読み込む
        print(f"入力ファイルを読み込んでいます: {input_path}")
        self.input_data = self.read_input_file(input_path)
        
        # 2. 出力用のワークブックを作成
        print("出力フォーマットを作成しています...")
        self.output_wb = self.create_output_workbook()
        
        # 3. データを変換して出力ワークブックに書き込む
        print("データを変換しています...")
        self.transform_data()
        
        # 4. ファイルを保存
        print(f"出力ファイルを保存しています: {output_path}")
        self.output_wb.save(output_path)
        print("変換完了！")
        
    def read_input_file(self, input_path):
        """
        入力ファイルを読み込む
        
        Parameters:
        -----------
        input_path : str
            入力ファイルのパス
            
        Returns:
        --------
        pandas.DataFrame
            読み込んだデータ
        """
        # ヘッダー行が7行目（0-indexedで6）にあるため、header=7で読み込む
        df = pd.read_excel(input_path, sheet_name='組立部品リスト', header=7)
        
        print(f"データを読み込みました: {len(df)} 行, {len(df.columns)} 列")
        return df
        
    def create_output_workbook(self):
        """
        出力用のワークブックを作成（output_format.xlsxに準拠）
        
        Returns:
        --------
        openpyxl.Workbook
            作成されたワークブック
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "変換結果"
        
        # 1-2行目のヘッダー（縦横結合）
        # A1-C2: Order名
        ws.merge_cells('A1:C2')
        ws.cell(row=1, column=1, value='Order名')
        
        # D1-F2: Order＃
        ws.merge_cells('D1:F2')
        ws.cell(row=1, column=4, value='Order＃')
        
        # G1-I2: Item＃
        ws.merge_cells('G1:I2')
        ws.cell(row=1, column=7, value='Item＃')
        
        # J1-P2: Item名称
        ws.merge_cells('J1:P2')
        ws.cell(row=1, column=10, value='Item名称')
        
        # Q1-T1: ｄ
        ws.merge_cells('Q1:T1')
        ws.cell(row=1, column=17, value='ｄ')
        
        # U1-X1: Ds
        ws.merge_cells('U1:X1')
        ws.cell(row=1, column=21, value='Ds')
        
        # Y1-AB1: Dm
        ws.merge_cells('Y1:AB1')
        ws.cell(row=1, column=25, value='Dm')
        
        # AC1-AF1: De
        ws.merge_cells('AC1:AF1')
        ws.cell(row=1, column=29, value='De')
        
        # AG1-AJ1: PD
        ws.merge_cells('AG1:AJ1')
        ws.cell(row=1, column=33, value='PD')
        
        # 2行目のサブヘッダー
        # Q2-R2: 部品数, S2-T2: 重量
        ws.merge_cells('Q2:R2')
        ws.cell(row=2, column=17, value='部品数')
        ws.merge_cells('S2:T2')
        ws.cell(row=2, column=19, value='重量')
        
        # U2-V2: 部品数, W2-X2: 重量
        ws.merge_cells('U2:V2')
        ws.cell(row=2, column=21, value='部品数')
        ws.merge_cells('W2:X2')
        ws.cell(row=2, column=23, value='重量')
        
        # Y2-Z2: 部品数, AA2-AB2: 重量
        ws.merge_cells('Y2:Z2')
        ws.cell(row=2, column=25, value='部品数')
        ws.merge_cells('AA2:AB2')
        ws.cell(row=2, column=27, value='重量')
        
        # AC2-AD2: 部品数, AE2-AF2: 重量
        ws.merge_cells('AC2:AD2')
        ws.cell(row=2, column=29, value='部品数')
        ws.merge_cells('AE2:AF2')
        ws.cell(row=2, column=31, value='重量')
        
        # AG2-AH2: 部品数, AI2-AJ2: 重量
        ws.merge_cells('AG2:AH2')
        ws.cell(row=2, column=33, value='部品数')
        ws.merge_cells('AI2:AJ2')
        ws.cell(row=2, column=35, value='重量')
        
        # ヘッダーのスタイルを設定
        self.apply_header_style(ws)
        
        return wb
        
    def apply_header_style(self, ws):
        """
        ヘッダー行にスタイルを適用
        
        Parameters:
        -----------
        ws : openpyxl.worksheet.worksheet.Worksheet
            ワークシート
        """
        # ヘッダー用のスタイル
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1行目と2行目にスタイルを適用
        for row in range(1, 3):
            for col in range(1, 36):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
        
    def transform_data(self):
        """
        データを変換して出力ワークブックに書き込む
        """
        ws = self.output_wb.active
        
        # ITEMごとにデータをグループ化して集計
        grouped_data = self.group_by_item()
        
        current_row = 3  # データは3行目から開始
        
        for item_key, item_data in grouped_data.items():
            try:
                # 基本情報を書き込む（結合セルの最初のセルに書き込む）
                # A列-C列: Order名
                ws.merge_cells(f'A{current_row}:C{current_row}')
                ws.cell(row=current_row, column=1, value=item_data['order_name'])
                
                # D列-F列: Order＃
                ws.merge_cells(f'D{current_row}:F{current_row}')
                ws.cell(row=current_row, column=4, value=item_data['order_num'])
                
                # G列-I列: Item＃
                ws.merge_cells(f'G{current_row}:I{current_row}')
                ws.cell(row=current_row, column=7, value=item_data['item_num'])
                
                # J列-P列: Item名称
                ws.merge_cells(f'J{current_row}:P{current_row}')
                ws.cell(row=current_row, column=10, value=item_data['item_name'])
                
                # カテゴリ別の部品数と重量を書き込む
                # d (Detail) - Q列:部品数, S列:重量
                ws.merge_cells(f'Q{current_row}:R{current_row}')
                ws.cell(row=current_row, column=17, value=item_data['d_count'])
                ws.merge_cells(f'S{current_row}:T{current_row}')
                ws.cell(row=current_row, column=19, value=round(item_data['d_weight'], 2))
                
                # Ds (Design start) - U列:部品数, W列:重量
                ws.merge_cells(f'U{current_row}:V{current_row}')
                ws.cell(row=current_row, column=21, value=item_data['ds_count'])
                ws.merge_cells(f'W{current_row}:X{current_row}')
                ws.cell(row=current_row, column=23, value=round(item_data['ds_weight'], 2))
                
                # Dm (Design middle) - Y列:部品数, AA列:重量
                ws.merge_cells(f'Y{current_row}:Z{current_row}')
                ws.cell(row=current_row, column=25, value=item_data['dm_count'])
                ws.merge_cells(f'AA{current_row}:AB{current_row}')
                ws.cell(row=current_row, column=27, value=round(item_data['dm_weight'], 2))
                
                # De (Design end) - AC列:部品数, AE列:重量
                ws.merge_cells(f'AC{current_row}:AD{current_row}')
                ws.cell(row=current_row, column=29, value=item_data['de_count'])
                ws.merge_cells(f'AE{current_row}:AF{current_row}')
                ws.cell(row=current_row, column=31, value=round(item_data['de_weight'], 2))
                
                # PD (Production drawing) - AG列:部品数, AI列:重量
                ws.merge_cells(f'AG{current_row}:AH{current_row}')
                ws.cell(row=current_row, column=33, value=item_data['pd_count'])
                ws.merge_cells(f'AI{current_row}:AJ{current_row}')
                ws.cell(row=current_row, column=35, value=round(item_data['pd_weight'], 2))
                
                current_row += 1
                
            except Exception as e:
                print(f"Item {item_key} でエラー: {str(e)}")
                continue
        
        # 列幅を調整
        for col in range(1, 36):
            ws.column_dimensions[get_column_letter(col)].width = 12
    
    def group_by_item(self):
        """
        ITEMごとにデータをグループ化して集計
        
        Returns:
        --------
        dict
            ITEM番号をキーとした辞書
        """
        grouped = {}
        
        # ITEMごとにグループ化
        for item_num in self.input_data['ITEM'].dropna().unique():
            item_rows = self.input_data[self.input_data['ITEM'] == item_num]
            
            if len(item_rows) == 0:
                continue
            
            # 最初の行から基本情報を取得
            first_row = item_rows.iloc[0]
            
            # 部品数と重量を集計
            total_count = len(item_rows)
            
            # 使用数量の合計
            usage_sum = item_rows['使用'].fillna(0).sum()
            
            # 単重の列（列番号22 = インデックス22）を取得して合計重量を計算
            unit_weight_col = item_rows.iloc[:, 22] if item_rows.shape[1] > 22 else pd.Series([0] * len(item_rows))
            total_weight = (item_rows['使用'].fillna(0) * unit_weight_col.fillna(0)).sum()
            
            # TODO: カテゴリ分類ロジックを実装
            # 現時点では全てPDカテゴリに振り分け
            item_data = {
                'order_name': 'TNPR',
                'order_num': '1021K457',
                'item_num': str(int(item_num)),
                'item_name': f'Item {int(item_num)}',
                'd_count': 0,
                'd_weight': 0.0,
                'ds_count': 0,
                'ds_weight': 0.0,
                'dm_count': 0,
                'dm_weight': 0.0,
                'de_count': 0,
                'de_weight': 0.0,
                'pd_count': total_count,
                'pd_weight': total_weight
            }
            
            grouped[int(item_num)] = item_data
        
        return grouped


# テスト用コード（このファイルを直接実行した場合のみ動作）
if __name__ == "__main__":
    converter = ExcelConverter()
    # テスト実行
    # converter.convert('test_input.xlsx', 'test_output.xlsx')
    print("converter.py が読み込まれました")
