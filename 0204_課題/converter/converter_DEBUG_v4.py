"""
Excel変換ロジック
入力ファイルを読み込み、指定のフォーマットに変換する
"""

import pandas as pd
import openpyxl
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelConverter:
    """Excelデータ変換クラス"""
    
    def __init__(self):
        """初期化"""
        self.input_data = None
        self.output_wb = None
    
    @staticmethod
    def classify_category(row, all_df):
        """
        フローチャートに基づく完全な分類ロジック
        
        Parameters:
        -----------
        row : pandas.Series
            処理対象の行
        all_df : pandas.DataFrame
            全データ（構成部材の参照用）
        
        Returns:
        --------
        str
            カテゴリ ('PD', 'De', 'Dm', 'Ds', 'd', None)
        """
        try:
            # データ取得
            parent_assy = row.get('PARENT ASSY NO')
            serial = row.get('SERIAL')
            
            # 使用数量
            usage = float(row.get('使用', 0)) if pd.notna(row.get('使用')) else 0.0
            
            # 単重（列22）
            unit_weight = 0.0
            if len(row) > 22 and pd.notna(row.iloc[22]):
                try:
                    unit_weight = float(row.iloc[22])
                except (ValueError, TypeError):
                    unit_weight = 0.0
            
            # 合計重量
            total_weight = usage * unit_weight
            
            # 材質（列9）
            material = str(row.iloc[9]).upper() if len(row) > 9 and pd.notna(row.iloc[9]) else ''
            
            # SUPPLY（列13）
            supply = str(row.iloc[13]).upper() if len(row) > 13 and pd.notna(row.iloc[13]) else ''
            
            # SUMMARY（列14）
            summary = str(row.iloc[14]).upper() if len(row) > 14 and pd.notna(row.iloc[14]) else ''
            
            # 部品名（列5）
            parts_name = str(row.iloc[5]).upper() if len(row) > 5 and pd.notna(row.iloc[5]) else ''
            
            # 【判定1】SUMMARY欄に MS- または JIS の記述があるか？
            if 'MS-' in summary or 'MS' in summary or 'JIS' in summary:
                return None  # 定義対象外
            
            # 【判定2】SUPPLY判定（ES, SU, ST, CS）
            if any(x in supply for x in ['ES', 'SU', 'ST', 'CS']):
                # SUPPLYに該当する記述がある場合
                
                # Name of parts に Roll または Roller が含まれるか？
                has_roll = 'ROLL' in parts_name or 'ROLLER' in parts_name
                
                if not has_roll:
                    # Roll/Rollerが含まれない → 定義対象外
                    return None
                
                # Roll/Rollerが含まれる場合、さらにBRG/Bearingをチェック
                has_brg = 'BRG' in parts_name or 'BEARING' in parts_name
                
                if has_brg:
                    # Roll/Roller かつ BRG/Bearing → 定義対象外
                    return None
                
                # Roll/Roller のみ（BRG/Bearingなし） → 次の判定へ進む
            
            # 【判定3】PARENT ASSY NOが空 → 対象外（この判定は後回し）
            # ※フローチャートでは、PARENT ASSY NOの判定はもっと後にある
            
            # 【判定3】部品(PARTS)単品の重量が 3000kg 以上か？
            if total_weight >= 3000:
                return 'Ds'
            
            # 【判定4-7】材質と重量による判定
            is_common_steel = any(x in material for x in ['SPCC', 'SPHC', 'SS'])
            
            if not is_common_steel:
                # 【判定5】一般鋼材以外で 400kg以上
                if total_weight >= 400:
                    return 'Ds'
            else:
                # 【判定6】構成部材にSPCC/SPHC/SS以外で300kg以上のものがあるか
                # 同じSERIALの他の部品を確認
                if pd.notna(serial):
                    same_serial_parts = all_df[all_df['SERIAL'] == serial]
                    for _, part in same_serial_parts.iterrows():
                        part_material = str(part.iloc[9]).upper() if len(part) > 9 and pd.notna(part.iloc[9]) else ''
                        part_usage = float(part.get('使用', 0)) if pd.notna(part.get('使用')) else 0.0
                        part_unit_weight = float(part.iloc[22]) if len(part) > 22 and pd.notna(part.iloc[22]) else 0.0
                        part_weight = part_usage * part_unit_weight
                        
                        # SPCC/SPHC/SS以外で300kg以上
                        if not any(x in part_material for x in ['SPCC', 'SPHC', 'SS']):
                            if part_weight >= 300:
                                return 'Ds'
            
            # 【判定7】SCM/SF材 + 300kg以上
            if 'SCM' in material or 'SF' in material:
                if total_weight >= 300:
                    return 'Ds'
            
            # 【判定8】非SPCC部品で100kg以上
            if not is_common_steel and total_weight >= 100:
                return 'Dm'
            
            # 【判定9-10】板厚80mm以上 + 300kg以上
            # ※板厚データが明確でないため、SIZE列から推測を試みる
            # SIZE列（列6）に厚み情報がある場合がある
            try:
                size_str = str(row.iloc[6]).upper() if len(row) > 6 and pd.notna(row.iloc[6]) else ''
                # 80mm以上の板厚を検出（例: "t80", "80mm", "PL80"など）
                import re
                thickness_match = re.search(r'(?:T|PL)?(\d+)(?:MM)?', size_str)
                if thickness_match:
                    thickness = int(thickness_match.group(1))
                    if thickness >= 80:
                        # 構成部材の重量が300kg以上か確認
                        if pd.notna(serial):
                            same_serial_parts = all_df[all_df['SERIAL'] == serial]
                            for _, part in same_serial_parts.iterrows():
                                part_usage = float(part.get('使用', 0)) if pd.notna(part.get('使用')) else 0.0
                                part_unit_weight = float(part.iloc[22]) if len(part) > 22 and pd.notna(part.iloc[22]) else 0.0
                                part_weight = part_usage * part_unit_weight
                                
                                if part_weight >= 300:
                                    return 'Ds'
            except:
                pass  # 板厚判定失敗時はスキップ
            
            # 【判定11】部品重量 ≥ 500kg
            if total_weight >= 500:
                return 'Dm'
            
            # 【判定12】PIPE/PIPING
            if 'PIPE' in parts_name or 'PIPING' in parts_name:
                return 'PD'
            
            # 【判定13】重量 < 50kg
            if total_weight < 50:
                return 'PD'
            
            # 【判定14】残部は後で振り分け（部品数で半分→重量で半分）
            return 'unclassified'
            
        except Exception as e:
            print(f"分類エラー: {str(e)}")
            return 'd'
        
    def convert(self, input_path, output_path):
        """
        メイン変換処理
        
        Parameters:
        -----------
        input_path : str
            入力ファイルのパス
        output_path : str
            出力ファイルのパス
        """
        # 1. 入力ファイルを読み込む
        print(f"入力ファイルを読み込んでいます: {input_path}")
        self.input_data = self.read_input_file(input_path)
        
        # 2. 出力用のワークブックを作成
        print("出力フォーマットを作成しています...")
        self.output_wb = self.create_output_workbook()
        
        # 3. データを変換して出力ワークブックに書き込む
        print("データを変換しています...")
        self.transform_data()
        
        # 4. ファイルを保存
        print(f"出力ファイルを保存しています: {output_path}")
        self.output_wb.save(output_path)
        print("変換完了！")
        
    def read_input_file(self, input_path):
        """
        入力ファイルを読み込む
        
        Parameters:
        -----------
        input_path : str
            入力ファイルのパス
            
        Returns:
        --------
        pandas.DataFrame
            読み込んだデータ
        """
        # ヘッダー行が7行目（0-indexedで6）にあるため、header=7で読み込む
        df = pd.read_excel(input_path, sheet_name='組立部品リスト', header=7)
        
        print(f"データを読み込みました: {len(df)} 行, {len(df.columns)} 列")
        return df
        
    def create_output_workbook(self):
        """
        出力用のワークブックを作成（output_format.xlsxに準拠）
        
        Returns:
        --------
        openpyxl.Workbook
            作成されたワークブック
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "変換結果"
        
        # 1-2行目のヘッダー（縦横結合）
        # A1-C2: Order名
        ws.merge_cells('A1:C2')
        ws.cell(row=1, column=1, value='Order名')
        
        # D1-F2: Order＃
        ws.merge_cells('D1:F2')
        ws.cell(row=1, column=4, value='Order＃')
        
        # G1-I2: Item＃
        ws.merge_cells('G1:I2')
        ws.cell(row=1, column=7, value='Item＃')
        
        # J1-P2: Item名称
        ws.merge_cells('J1:P2')
        ws.cell(row=1, column=10, value='Item名称')
        
        # Q1-T1: ｄ
        ws.merge_cells('Q1:T1')
        ws.cell(row=1, column=17, value='ｄ')
        
        # U1-X1: Ds
        ws.merge_cells('U1:X1')
        ws.cell(row=1, column=21, value='Ds')
        
        # Y1-AB1: Dm
        ws.merge_cells('Y1:AB1')
        ws.cell(row=1, column=25, value='Dm')
        
        # AC1-AF1: De
        ws.merge_cells('AC1:AF1')
        ws.cell(row=1, column=29, value='De')
        
        # AG1-AJ1: PD
        ws.merge_cells('AG1:AJ1')
        ws.cell(row=1, column=33, value='PD')
        
        # 2行目のサブヘッダー
        # Q2-R2: 部品数, S2-T2: 重量
        ws.merge_cells('Q2:R2')
        ws.cell(row=2, column=17, value='部品数')
        ws.merge_cells('S2:T2')
        ws.cell(row=2, column=19, value='重量')
        
        # U2-V2: 部品数, W2-X2: 重量
        ws.merge_cells('U2:V2')
        ws.cell(row=2, column=21, value='部品数')
        ws.merge_cells('W2:X2')
        ws.cell(row=2, column=23, value='重量')
        
        # Y2-Z2: 部品数, AA2-AB2: 重量
        ws.merge_cells('Y2:Z2')
        ws.cell(row=2, column=25, value='部品数')
        ws.merge_cells('AA2:AB2')
        ws.cell(row=2, column=27, value='重量')
        
        # AC2-AD2: 部品数, AE2-AF2: 重量
        ws.merge_cells('AC2:AD2')
        ws.cell(row=2, column=29, value='部品数')
        ws.merge_cells('AE2:AF2')
        ws.cell(row=2, column=31, value='重量')
        
        # AG2-AH2: 部品数, AI2-AJ2: 重量
        ws.merge_cells('AG2:AH2')
        ws.cell(row=2, column=33, value='部品数')
        ws.merge_cells('AI2:AJ2')
        ws.cell(row=2, column=35, value='重量')
        
        # ヘッダーのスタイルを設定
        self.apply_header_style(ws)
        
        return wb
        
    def apply_header_style(self, ws):
        """
        ヘッダー行にスタイルを適用（フォントと罫線のみ）
        
        Parameters:
        -----------
        ws : openpyxl.worksheet.worksheet.Worksheet
            ワークシート
        """
        # ヘッダー用のフォント（太字）
        header_font = Font(bold=True, size=11)
        
        # 1-2行目に太字を適用
        for row in range(1, 3):
            for col in range(1, 37):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
    
    def apply_data_style(self, ws, max_row):
        """
        データ行にスタイルを適用（罫線と背景色）
        
        Parameters:
        -----------
        ws : openpyxl.worksheet.worksheet.Worksheet
            ワークシート
        max_row : int
            最大行数
        """
        # 罫線
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 中央揃え
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # 背景色の定義
        # A-P列: 白
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        # Q-T列: 薄い黄色
        yellow_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
        
        # U-X列: 薄い赤
        red_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
        
        # Y-AB列: 薄いオレンジ
        orange_fill = PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid')
        
        # AC-AF列: 薄い緑
        green_fill = PatternFill(start_color='E0FFE0', end_color='E0FFE0', fill_type='solid')
        
        # AG-AJ列: 薄いグレー
        gray_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        
        # 全ての行にスタイルを適用
        for row in range(1, max_row + 1):
            for col in range(1, 37):  # A-AJ列（36列）
                cell = ws.cell(row=row, column=col)
                
                # 罫線を適用
                cell.border = border
                
                # 中央揃え
                cell.alignment = center_alignment
                
                # 背景色を適用（全ての行）
                if 1 <= col <= 16:  # A-P列
                    cell.fill = white_fill
                elif 17 <= col <= 20:  # Q-T列 (ｄ)
                    cell.fill = yellow_fill
                elif 21 <= col <= 24:  # U-X列 (Ds)
                    cell.fill = red_fill
                elif 25 <= col <= 28:  # Y-AB列 (Dm)
                    cell.fill = orange_fill
                elif 29 <= col <= 32:  # AC-AF列 (De)
                    cell.fill = green_fill
                elif 33 <= col <= 36:  # AG-AJ列 (PD)
                    cell.fill = gray_fill
        
    def transform_data(self):
        """
        データを変換して出力ワークブックに書き込む
        """
        ws = self.output_wb.active
        
        # ITEMごとにデータをグループ化して集計
        grouped_data = self.group_by_item()
        
        current_row = 3  # データは3行目から開始
        
        for item_key, item_data in grouped_data.items():
            try:
                # 基本情報を書き込む（結合セルの最初のセルに書き込む）
                # A列-C列: Order名
                ws.merge_cells(f'A{current_row}:C{current_row}')
                ws.cell(row=current_row, column=1, value=item_data['order_name'])
                
                # D列-F列: Order＃
                ws.merge_cells(f'D{current_row}:F{current_row}')
                ws.cell(row=current_row, column=4, value=item_data['order_num'])
                
                # G列-I列: Item＃
                ws.merge_cells(f'G{current_row}:I{current_row}')
                ws.cell(row=current_row, column=7, value=item_data['item_num'])
                
                # J列-P列: Item名称
                ws.merge_cells(f'J{current_row}:P{current_row}')
                ws.cell(row=current_row, column=10, value=item_data['item_name'])
                
                # カテゴリ別の部品数と重量を書き込む
                # d (Detail) - Q列:部品数, S列:重量
                ws.merge_cells(f'Q{current_row}:R{current_row}')
                ws.cell(row=current_row, column=17, value=item_data['d_count'])
                ws.merge_cells(f'S{current_row}:T{current_row}')
                ws.cell(row=current_row, column=19, value=round(item_data['d_weight'], 2))
                
                # Ds (Design start) - U列:部品数, W列:重量
                ws.merge_cells(f'U{current_row}:V{current_row}')
                ws.cell(row=current_row, column=21, value=item_data['ds_count'])
                ws.merge_cells(f'W{current_row}:X{current_row}')
                ws.cell(row=current_row, column=23, value=round(item_data['ds_weight'], 2))
                
                # Dm (Design middle) - Y列:部品数, AA列:重量
                ws.merge_cells(f'Y{current_row}:Z{current_row}')
                ws.cell(row=current_row, column=25, value=item_data['dm_count'])
                ws.merge_cells(f'AA{current_row}:AB{current_row}')
                ws.cell(row=current_row, column=27, value=round(item_data['dm_weight'], 2))
                
                # De (Design end) - AC列:部品数, AE列:重量
                ws.merge_cells(f'AC{current_row}:AD{current_row}')
                ws.cell(row=current_row, column=29, value=item_data['de_count'])
                ws.merge_cells(f'AE{current_row}:AF{current_row}')
                ws.cell(row=current_row, column=31, value=round(item_data['de_weight'], 2))
                
                # PD (Production drawing) - AG列:部品数, AI列:重量
                ws.merge_cells(f'AG{current_row}:AH{current_row}')
                ws.cell(row=current_row, column=33, value=item_data['pd_count'])
                ws.merge_cells(f'AI{current_row}:AJ{current_row}')
                ws.cell(row=current_row, column=35, value=round(item_data['pd_weight'], 2))
                
                current_row += 1
                
            except Exception as e:
                print(f"Item {item_key} でエラー: {str(e)}")
                continue
        
        # 列幅を調整
        for col in range(1, 36):
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        # データ行にスタイルを適用（罫線と背景色）
        self.apply_data_style(ws, current_row - 1)
    
    def group_by_item(self):
        """
        ITEMごとにデータをグループ化して集計
        PARTS番号が連番の場合は1つとしてカウント
        
        Returns:
        --------
        dict
            ITEM番号をキーとした辞書
        """
        grouped = {}
        
        # ITEMごとにグループ化
        for item_num in self.input_data['ITEM'].dropna().unique():
            try:
                item_rows = self.input_data[self.input_data['ITEM'] == item_num]
                
                if len(item_rows) == 0:
                    continue
                
                # 最初の行から基本情報を取得
                first_row = item_rows.iloc[0]
                
                # SERIALごとに、さらにPARTSでグループ化
                serial_groups = {}
                
                for serial in item_rows['SERIAL'].dropna().unique():
                    serial_rows = item_rows[item_rows['SERIAL'] == serial]
                    
                    # PARTSごとにグループ化（PARTS番号でまとめる）
                    parts_list = []
                    for parts_num in serial_rows.iloc[:, 3].dropna().unique():  # 列3 = PARTS
                        parts_rows = serial_rows[serial_rows.iloc[:, 3] == parts_num]
                        
                        # PARTSグループ全体の重量を計算
                        total_weight = 0.0
                        for _, row in parts_rows.iterrows():
                            usage = float(row.get('使用', 0)) if pd.notna(row.get('使用')) else 0.0
                            unit_weight = float(row.iloc[22]) if len(row) > 22 and pd.notna(row.iloc[22]) else 0.0
                            total_weight += usage * unit_weight
                        
                        # PARTSグループを1つの部品として扱う
                        parts_list.append({
                            'parts_num': parts_num,
                            'weight': total_weight,
                            'rows': parts_rows,
                            'representative_row': parts_rows.iloc[0]  # 代表行
                        })
                    
                    serial_groups[serial] = parts_list
                
                # カテゴリ別に分類して集計
                category_stats = {
                    'd': {'count': 0, 'weight': 0.0, 'parts': []},
                    'Ds': {'count': 0, 'weight': 0.0, 'parts': []},
                    'Dm': {'count': 0, 'weight': 0.0, 'parts': []},
                    'De': {'count': 0, 'weight': 0.0, 'parts': []},
                    'PD': {'count': 0, 'weight': 0.0, 'parts': []},
                    'unclassified': {'count': 0, 'weight': 0.0, 'parts': []}
                }
                
                # 各PARTSグループをカテゴリ分類
                classification_debug = []
                for serial, parts_list in serial_groups.items():
                    for parts_group in parts_list:
                        try:
                            # PARTSグループの重量を使って判定
                            # 代表行を使うが、重量はグループ全体の重量を使用
                            rep_row = parts_group['representative_row'].copy()
                            group_weight = parts_group['weight']
                            
                            # SUMMARYチェック（デバッグ）
                            summary = str(rep_row.iloc[14]).upper() if len(rep_row) > 14 and pd.notna(rep_row.iloc[14]) else ''
                            has_ms_jis = 'MS-' in summary or 'MS' in summary or 'JIS' in summary
                            
                            # 判定用に、グループ全体の重量を反映した仮の行を作成
                            # 使用数量と単重を調整して、total_weightがgroup_weightになるようにする
                            if group_weight > 0:
                                rep_row.iloc[7] = 1.0  # 使用数量を1に
                                rep_row.iloc[22] = group_weight  # 単重をグループ重量に
                            
                            # カテゴリを判定
                            category = self.classify_category(rep_row, self.input_data)
                            
                            # デバッグ情報
                            parts_num = parts_group['parts_num']
                            classification_debug.append({
                                'serial': serial,
                                'parts': parts_num,
                                'weight': group_weight,
                                'has_ms': has_ms_jis,
                                'summary': summary[:30],
                                'category': category
                            })
                            
                            if category is None:
                                continue  # 対象外
                            
                            # PARTSグループを1つとしてカウント
                            category_stats[category]['count'] += 1
                            category_stats[category]['weight'] += group_weight
                            category_stats[category]['parts'].append({
                                'weight': group_weight,
                                'parts_num': parts_group['parts_num']
                            })
                            
                        except Exception as e:
                            print(f"PARTS {parts_group['parts_num']} の処理中にエラー: {str(e)}")
                            continue
                
                # デバッグ出力（最初の20個）
                if len(classification_debug) > 0:
                    print(f"\n=== ITEM {item_num} の分類デバッグ（最初の20個）===")
                    for i, info in enumerate(classification_debug[:20]):
                        print(f"{info['serial']}-{info['parts']}: {info['weight']:.2f}kg, MS={info['has_ms']}, {info['summary']} → {info['category']}")
                    
                    # PD分類された部品の詳細（最初の20個）
                    pd_parts = [info for info in classification_debug if info['category'] == 'PD']
                    if len(pd_parts) > 0:
                        print(f"\n--- PD分類された部品（最初の20個）---")
                        for i, info in enumerate(pd_parts[:20]):
                            print(f"{info['serial']}-{info['parts']}: {info['weight']:.2f}kg, MS={info['has_ms']}")
                        
                        # 50kg以上のPD部品を詳細表示
                        heavy_pd = [info for info in pd_parts if info['weight'] >= 50]
                        if len(heavy_pd) > 0:
                            print(f"\n!!! 50kg以上なのにPDになっている部品 !!!")
                            for info in heavy_pd:
                                print(f"{info['serial']}-{info['parts']}: {info['weight']:.2f}kg, SUMMARY={info['summary']}")
                        
                        # 重量分布
                        weights = [info['weight'] for info in pd_parts]
                        under_50 = sum(1 for w in weights if w < 50)
                        over_50 = sum(1 for w in weights if w >= 50)
                        print(f"PD部品の重量分布: <50kg={under_50}個, ≥50kg={over_50}個")
                    
                    print(f"総PARTS数: {len(classification_debug)}個")
                    
                    # カテゴリ別集計も表示
                    cat_count = {}
                    for info in classification_debug:
                        cat = info['category'] if info['category'] else 'None'
                        cat_count[cat] = cat_count.get(cat, 0) + 1
                    print(f"カテゴリ別: {cat_count}")
                
                # 【判定12】未分類の部品を振り分け
                unclassified_parts = category_stats['unclassified']['parts']
                if len(unclassified_parts) > 0:
                    # 部品を重量順にソート
                    unclassified_parts.sort(key=lambda x: x['weight'], reverse=True)
                    
                    # 部品数で半分に分ける
                    half_count = len(unclassified_parts) // 2
                    
                    # 重い方の半分 → Dm
                    for i in range(half_count):
                        part = unclassified_parts[i]
                        category_stats['Dm']['count'] += 1
                        category_stats['Dm']['weight'] += part['weight']
                    
                    # 軽い方の半分 → De
                    for i in range(half_count, len(unclassified_parts)):
                        part = unclassified_parts[i]
                        category_stats['De']['count'] += 1
                        category_stats['De']['weight'] += part['weight']
                
                # Item名称を取得
                item_name = f"Item {int(item_num)}"
                try:
                    if len(first_row) > 5 and pd.notna(first_row.iloc[5]):
                        item_name_col5 = str(first_row.iloc[5]).strip()
                        if item_name_col5:
                            item_name = item_name_col5[:50]  # 最大50文字
                except Exception as e:
                    print(f"Item名取得エラー: {str(e)}")
                
                # 結果を格納
                item_data = {
                    'order_name': 'TNPR',
                    'order_num': '1021K457',
                    'item_num': str(int(item_num)),
                    'item_name': item_name,
                    'd_count': category_stats['d']['count'],
                    'd_weight': category_stats['d']['weight'],
                    'ds_count': category_stats['Ds']['count'],
                    'ds_weight': category_stats['Ds']['weight'],
                    'dm_count': category_stats['Dm']['count'],
                    'dm_weight': category_stats['Dm']['weight'],
                    'de_count': category_stats['De']['count'],
                    'de_weight': category_stats['De']['weight'],
                    'pd_count': category_stats['PD']['count'],
                    'pd_weight': category_stats['PD']['weight']
                }
                
                grouped[int(item_num)] = item_data
                
            except Exception as e:
                print(f"ITEM {item_num} の処理中にエラー: {str(e)}")
                import traceback
                traceback.print_exc()
                continue
        
        return grouped


# テスト用コード（このファイルを直接実行した場合のみ動作）
if __name__ == "__main__":
    converter = ExcelConverter()
    # テスト実行
    # converter.convert('test_input.xlsx', 'test_output.xlsx')
    print("converter.py が読み込まれました")
